"""
Exportação para XLSX — Simulador de Financiamento Imobiliário
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from typing import List
from engine import ResultadoCenario, AnaliseInvestimento

# Paleta
AZUL_ESCURO  = "1F4E79"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "BDD7EE"
VERDE        = "375623"
VERDE_CLARO  = "E2EFDA"
AMARELO      = "FFD966"
CINZA        = "D6DCE4"
LARANJA      = "833C00"
BRANCO       = "FFFFFF"
CINZA_CLARO  = "F2F2F2"

BRL  = '"R$" #,##0.00'
PCT  = '0.00%'
NUM  = '#,##0'
MULT = '0.00"x"'

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, cols_vals: list, bg=AZUL_ESCURO, fg=BRANCO, bold=True, size=10):
    for col, val in enumerate(cols_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = Font(bold=bold, color=fg, size=size)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _border()

def _data_row(ws, row, cols_vals: list, bg=BRANCO, bold=False, num_fmts: dict = None):
    for col, val in enumerate(cols_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = Font(bold=bold, color="000000", size=9)
        c.alignment = Alignment(horizontal='right' if isinstance(val, (int, float)) else 'left', vertical='center')
        c.border = _border()
        if num_fmts and col in num_fmts:
            c.number_format = num_fmts[col]


def exportar_xlsx(resultados: List[ResultadoCenario]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _aba_resumo(wb, resultados)
    for res in resultados:
        _aba_fluxo(wb, res)
    _aba_analise(wb, resultados)
    _aba_indicadores(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _aba_resumo(wb: Workbook, resultados: List[ResultadoCenario]):
    ws = wb.create_sheet("Resumo")
    ws.sheet_properties.tabColor = AZUL_ESCURO

    ws.column_dimensions['A'].width = 32
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 22

    # Título
    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = "RESUMO COMPARATIVO — SIMULADOR DE FINANCIAMENTO IMOBILIÁRIO"
    c.fill = _fill(AZUL_ESCURO); c.font = _font(bold=True, size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header cenários
    hdrs = ['MÉTRICA'] + [f"CENÁRIO {i+1}\n{r.premissas.nome}" for i, r in enumerate(resultados)]
    _header_row(ws, 2, hdrs, bg=AZUL_MEDIO, size=10)
    ws.row_dimensions[2].height = 28

    secoes = [
        ("PREMISSAS", [
            ("Valor do Imóvel",       lambda r: r.premissas.valor_imovel,          BRL),
            ("Entrada",               lambda r: r.premissas.entrada,               BRL),
            ("Valor Financiado",      lambda r: r.premissas.valor_financiado,      BRL),
            ("Taxa Efetiva a.a.",     lambda r: r.premissas.taxa_efetiva_aa/100,   PCT),
            ("Taxa Mensal",           lambda r: r.premissas.taxa_mensal,           '0.0000%'),
            ("Prazo Contratual",      lambda r: r.premissas.prazo_meses,           NUM),
            ("Sistema",               lambda r: r.premissas.sistema,               '@'),
            ("PMT após balão",        lambda r: r.premissas.pmt_apos_balao,        '@'),
            ("Nº de Balões",          lambda r: len(r.premissas.baloes),           NUM),
        ]),
        ("RESULTADOS", [
            ("PMT Inicial",           lambda r: r.pmt_inicial,                     BRL),
            ("PMT Final",             lambda r: r.pmt_final,                       BRL),
            ("Prazo Efetivo (meses)", lambda r: r.prazo_efetivo,                   NUM),
            ("Meses Economizados",    lambda r: r.premissas.prazo_meses - r.prazo_efetivo, NUM),
        ]),
        ("TOTAIS", [
            ("Total PMTs",            lambda r: r.total_pmt,                       BRL),
            ("Total Balões",          lambda r: r.total_balao,                     BRL),
            ("Total Juros",           lambda r: r.total_juros,                     BRL),
            ("Total Amortizado",      lambda r: r.total_amortizacao,               BRL),
            ("Entrada",               lambda r: r.premissas.entrada,               BRL),
            ("TOTAL PAGO",            lambda r: r.total_pago,                      BRL),
            ("Multiplicador",         lambda r: r.multiplicador,                   MULT),
        ]),
        ("ECONOMIA vs CENÁRIO 1", [
            ("Economia Total (R$)",   lambda r: resultados[0].total_pago - r.total_pago,   BRL),
            ("Economia Juros (R$)",   lambda r: resultados[0].total_juros - r.total_juros, BRL),
            ("Meses a menos",         lambda r: resultados[0].prazo_efetivo - r.prazo_efetivo, NUM),
        ]),
    ]

    row = 3
    for sec_nome, metricas in secoes:
        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1, value=sec_nome)
        c.fill = _fill(AZUL_MEDIO); c.font = _font(bold=True, color=BRANCO, size=10)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        for i, (label, fn, fmt) in enumerate(metricas):
            is_total = label == "TOTAL PAGO"
            bg = AMARELO if is_total else (CINZA_CLARO if i % 2 == 0 else BRANCO)
            vals = [label] + [fn(r) for r in resultados]
            fmts = {j+2: fmt for j in range(len(resultados))}
            _data_row(ws, row, vals, bg=bg, bold=is_total, num_fmts=fmts)
            # Label em azul claro
            ws.cell(row=row, column=1).fill = _fill("D6E4F0")
            ws.cell(row=row, column=1).font = Font(bold=True, size=9)
            ws.row_dimensions[row].height = 18
            row += 1

    # Colorir economia negativa em vermelho
    # (simplificado — aplicar formatação condicional manualmente se necessário)


def _aba_fluxo(wb: Workbook, res: ResultadoCenario):
    p  = res.premissas
    ws = wb.create_sheet(f"Fluxo {p.nome[:10]}")
    ws.sheet_properties.tabColor = AZUL_MEDIO

    # Larguras
    widths = [8, 12, 16, 16, 16, 16, 16, 16, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = f"{p.nome} — {p.sistema} | {p.pmt_apos_balao} | {p.taxa_efetiva_aa:.2f}% a.a. | R${p.valor_financiado/1e6:.2f}M | {p.prazo_meses}m"
    c.fill = _fill(AZUL_ESCURO); c.font = _font(bold=True, size=11)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    # Subtítulo
    ws.merge_cells('A2:J2')
    c = ws['A2']
    c.value = f"Entrada R${p.entrada/1e6:.2f}M | {len(p.baloes)} balões | PMT inicial: R${res.pmt_inicial:,.2f} | Prazo efetivo: {res.prazo_efetivo} meses"
    c.fill = _fill(AZUL_CLARO); c.font = Font(italic=True, color=AZUL_ESCURO, size=9)
    c.alignment = Alignment(horizontal='center')

    hdrs = ['Mês','Saldo Inicial','Juros','Amortização','PMT','Balão','Correção','Saldo Final','Evento','PMT Novo']
    _header_row(ws, 3, hdrs, bg=AZUL_MEDIO)
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = 'A4'

    fmts = {2: BRL, 3: BRL, 4: BRL, 5: BRL, 6: BRL, 7: BRL, 8: BRL}

    for i, linha in enumerate(res.fluxo):
        row = i + 4
        is_balao = linha.balao > 0
        bg = "FFF2CC" if is_balao else (CINZA_CLARO if i % 2 == 0 else BRANCO)
        vals = [
            linha.mes, linha.saldo_inicial, linha.juros, linha.amortizacao,
            linha.pmt, linha.balao, linha.correcao_saldo, linha.saldo_final,
            linha.evento, linha.pmt_novo or ''
        ]
        _data_row(ws, row, vals, bg=bg, num_fmts=fmts)
        if is_balao:
            ws.cell(row=row, column=9).font = Font(bold=True, color="7F4F00", size=9)
        ws.row_dimensions[row].height = 16

    # Totais
    tot_row = len(res.fluxo) + 4
    _header_row(ws, tot_row, ['TOTAIS','','',f'R${res.total_amortizacao:,.2f}',
                               f'R${res.total_pmt:,.2f}',f'R${res.total_balao:,.2f}','',
                               f'TOTAL PAGO: R${res.total_pago:,.2f}','',''],
                bg=CINZA, fg="000000", bold=True)


def _aba_analise(wb: Workbook, resultados: List[ResultadoCenario]):
    ws = wb.create_sheet("Análise Invest.")
    ws.sheet_properties.tabColor = VERDE
    ws.column_dimensions['A'].width = 35
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 22

    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = "ANÁLISE DE INVESTIMENTO — VPL | TIR | CUSTO DE OPORTUNIDADE | RENT VS BUY"
    c.fill = _fill(VERDE); c.font = _font(bold=True, size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    _header_row(ws, 2, ['MÉTRICA'] + [f"C{i+1} — {r.premissas.nome}" for i, r in enumerate(resultados)],
                bg=AZUL_MEDIO)

    row = 3
    for i, res in enumerate(resultados):
        ai = AnaliseInvestimento(res)
        co = ai.custo_oportunidade()
        tir = ai.tir_imovel()
        vpl = ai.vpl_financiamento()
        be  = ai.ponto_equilibrio()

        # Seção por cenário
        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1, value=f"CENÁRIO {i+1} — {res.premissas.nome}")
        c.fill = _fill(AZUL_MEDIO); c.font = _font(bold=True, size=10)
        ws.row_dimensions[row].height = 20
        row += 1

        metricas = [
            ("VPL do Financiamento (R$)",        vpl,                              BRL),
            ("TIR do Imóvel (% a.a.)",            (tir or 0) / 100,                PCT),
            ("Cap Rate Bruto (% a.a.)",           ai.cap_rate() / 100,             PCT),
            ("Yield Líquido (% a.a.)",            ai.yield_liquido() / 100,        PCT),
            ("Break-even Valorização (% a.a.)",   ai.break_even_valorizacao()/100, PCT),
            ("Ponto Equilíbrio Rent vs Buy",      f"Mês {be}" if be else "Não atingido", '@'),
            ("Patrimônio — Imóvel Final (R$)",    co['valor_imovel_final'],         BRL),
            ("Patrimônio — Renda Fixa (R$)",      co['patrimonio_renda_fixa'],      BRL),
            ("Vantagem do Imóvel (R$)",           co['vantagem_imovel'],            BRL),
            ("Total Investido no Financiamento",  co['total_investido_financiamento'], BRL),
        ]

        for j, (label, val, fmt) in enumerate(metricas):
            bg = VERDE_CLARO if j % 2 == 0 else BRANCO
            c_label = ws.cell(row=row, column=1, value=label)
            c_label.fill = _fill("D6E4F0"); c_label.font = Font(bold=True, size=9)
            c_label.border = _border()

            c_val = ws.cell(row=row, column=i+2, value=val)
            c_val.fill = _fill(bg); c_val.font = Font(size=9, bold=True)
            c_val.alignment = Alignment(horizontal='right')
            c_val.border = _border()
            if fmt != '@':
                c_val.number_format = fmt

            # Vantagem em verde/vermelho
            if label == "Vantagem do Imóvel (R$)" and isinstance(val, (int, float)):
                c_val.font = Font(size=9, bold=True, color="375623" if val >= 0 else "C00000")

            ws.row_dimensions[row].height = 18
            row += 1

        row += 1  # espaço entre cenários


def _aba_indicadores(wb: Workbook):
    from indicadores import get_dataframe, resumo_indicadores
    ws = wb.create_sheet("Indicadores")
    ws.sheet_properties.tabColor = LARANJA

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = "DATABASE DE INDICADORES ECONÔMICOS — Taxas mensais (%)"
    c.fill = _fill(LARANJA); c.font = _font(bold=True, size=12)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Resumo primeiro
    df_res = resumo_indicadores()
    _header_row(ws, 2, list(df_res.columns), bg="C55A11")
    for i, row_data in df_res.iterrows():
        r = i + 3
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill = _fill(CINZA_CLARO if i % 2 == 0 else BRANCO)
            c.font = Font(size=9)
            c.border = _border()
            c.alignment = Alignment(horizontal='right' if isinstance(val, float) else 'left')
        ws.row_dimensions[r].height = 16

    # Histórico completo abaixo
    start = len(df_res) + 5
    ws.merge_cells(f'A{start}:I{start}')
    c = ws.cell(row=start, column=1, value="HISTÓRICO MENSAL (2015–2025)")
    c.fill = _fill(LARANJA); c.font = _font(bold=True)
    c.alignment = Alignment(horizontal='center')

    df_hist = get_dataframe()[['ano','mes','IPCA','IGP-M','INCC','CUB/PR','SELIC','TR','Poupança']]
    hdrs = list(df_hist.columns)
    _header_row(ws, start+1, hdrs, bg="C55A11")

    for i, row_data in df_hist.iterrows():
        r = start + 2 + i
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill = _fill(CINZA_CLARO if i % 2 == 0 else BRANCO)
            c.font = Font(size=9)
            c.border = _border()
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 14

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14
